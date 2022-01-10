# NAME: DIKUBA KINDJEL ALICIA ASHLEY
#EMAIL: dikuba.ashley@ictuniveristy.edu.cm
#MATRICULATE:ICTU20201109
from openpyxl import workbook,load_workbook
wb = load_workbook('employeedata.xlsx')
#
ws= wb.active
range = ws["B2":"B31"]
def employeedatabase():
  for cell in range:
    for x in cell:
        print(x.value)
employeedatabase()
print("*** Changed***")

def updated_employeedatabase():
  for cell in range:
    for x in cell:

      text =x.value

      changesufix = text.replace("helpinghands.cm","handsinhands.org")
      x.value = changesufix
    print(changesufix)
    wb.save('updatedemployeedata.xlsx')
updated_employeedatabase()

#***working***on***CSV***

text = open("employeedataSheet1.csv","r")

#updating the column data from
text = ''.join([k for k in text])
text = text.replace('helpinghands.cm','handsinhand.org')
x = open("updatedemployeedata", "w")
x.writelines(text)
x.close()

